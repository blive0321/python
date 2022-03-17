import openpyxl
import time
from openpyxl import styles
from openpyxl.styles import Alignment, Border, Side, PatternFill
# Alignment 對齊
# Border, Side 框線
# PatternFill Cell背景色

taipeiTime = time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())

workbook = openpyxl.Workbook()

# ----------- 加入資料到當前work sheet工作表 -----------
data = [["Date","config","test item","result","logs","remark"]]
ws1 = workbook.active                 # 用active指定上一次關閉檔案前的當前工作表
ws1.title = "Sample-brandon"
for row in data:
    ws1.append(row)

writeData = [taipeiTime,"config3-1","power cycles","PASS","nothing","UCR"]
ws1.cell(row=2, column=1, value=writeData[0])
ws1.cell(row=2, column=2, value=writeData[1])
ws1.cell(row=2, column=3, value=writeData[2])
ws1.cell(row=2, column=4, value=writeData[3])
ws1.cell(row=2, column=5, value=writeData[4])
ws1.cell(row=2, column=6, value=writeData[5])

# ----------- 調整列、行的寬和高 -----------
ws1.column_dimensions['A'].width = 20.0
ws1.row_dimensions[1].height = 15.0
ws1.cell(1,6).font = styles.Font(name='Calibri', size=14, color='ff9300')

# ----------- 對齊 Alignment -----------
ws1["A1"].alignment = Alignment(horizontal='center')
ws1["A3"] = "OpenPyXL"
ws1["A3"].alignment = Alignment(text_rotation = 90)

# ----------- 調整框線 Border, Side -----------
pink = "00FF00FF"
thick = Side(border_style="thick", color=pink)
ws1["A1"].border = Border(top=thick, left=thick, right=thick, bottom=thick)

# ----------- 填滿cell的背景顏色 PatternFill -----------
yellow = "ffff00"
for rows in ws1.iter_rows(min_row=1, max_row=10, min_col=1, max_col=12):
    for cell in rows:
        if cell.row % 2:
            cell.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type = "solid")

# ----------- 合併儲存格 merge cells -----------
ws1.merge_cells("A5:L5")
ws1["A5"] = "Merge Cells"
ws1["A5"].alignment = Alignment(horizontal='center')

# ----------- 加入新的work sheet 工作表 -----------
ws2 = workbook.create_sheet("sheet_2")
ws2.sheet_properties.tabColor = "1072BA"
ws3 = workbook.create_sheet("sheet_3")
ws3.sheet_properties.tabColor = "1072BA"

# ----------- 加入資料到別的工作表 -----------
data = [["Config","fw-util","fruit-util","L10 S/N","SB S/N"]]
ws2 = workbook["sheet_2"]
for row in data:
    ws2.append(row)

# ----------- 儲存檔案 -----------
workbook.save("test.xlsx")
